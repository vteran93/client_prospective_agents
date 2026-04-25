"""
tools/excel_tool.py — Export qualified leads to a formatted Excel workbook.

Sheets:
  1. HOT     — Tier 1 leads (red header)
  2. WARM    — Tier 2 leads (orange header)
  3. COLD    — Tier 3 leads (blue header)
  4. TODOS   — All leads combined
  5. RESUMEN — Campaign summary / RunReport

Column groups (applied to HOT / WARM / COLD / TODOS):
  Contact | Location | Ratings | Digital | Profile | Timing | Score

Color coding rows by tier:
  HOT  → pastel red   #FFCCCC
  WARM → pastel orange #FFE5B4
  COLD → pastel blue   #CCE5FF
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Sequence

from models import QualifiedLead, RoutePlan, RunReport

# Conditional import — openpyxl must be installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    _OPENPYXL_OK = True
except ImportError:  # pragma: no cover
    _OPENPYXL_OK = False


# ──────────────────────────────────────────────────────────────────
# Colour palette
# ──────────────────────────────────────────────────────────────────

_TIER_ROW_FILL = {
    "HOT": PatternFill("solid", fgColor="FFCCCC") if _OPENPYXL_OK else None,
    "WARM": PatternFill("solid", fgColor="FFE5B4") if _OPENPYXL_OK else None,
    "COLD": PatternFill("solid", fgColor="CCE5FF") if _OPENPYXL_OK else None,
}
_HEADER_FILL = {
    "HOT": PatternFill("solid", fgColor="C00000") if _OPENPYXL_OK else None,
    "WARM": PatternFill("solid", fgColor="E26B0A") if _OPENPYXL_OK else None,
    "COLD": PatternFill("solid", fgColor="1F4E79") if _OPENPYXL_OK else None,
    "TODOS": PatternFill("solid", fgColor="2E4057") if _OPENPYXL_OK else None,
    "RESUMEN": PatternFill("solid", fgColor="375623") if _OPENPYXL_OK else None,
    "RUTA": PatternFill("solid", fgColor="BDD7EE") if _OPENPYXL_OK else None,
}
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10) if _OPENPYXL_OK else None
_ROUTE_HEADER_FONT = Font(bold=True, color="000000", size=10) if _OPENPYXL_OK else None
_THIN = Side(style="thin") if _OPENPYXL_OK else None
_BORDER = (
    Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN) if _OPENPYXL_OK else None
)
_ROUTE_TIER_FILL = {
    "HOT": PatternFill("solid", fgColor="C6EFCE") if _OPENPYXL_OK else None,
    "WARM": PatternFill("solid", fgColor="FFEB9C") if _OPENPYXL_OK else None,
    "COLD": PatternFill("solid", fgColor="D9D9D9") if _OPENPYXL_OK else None,
}

# ──────────────────────────────────────────────────────────────────
# Column definitions  (header, field_extractor)
# ──────────────────────────────────────────────────────────────────


def _get(lead: QualifiedLead, *attrs):
    obj = lead
    for attr in attrs:
        obj = getattr(obj, attr, None)
        if obj is None:
            return ""
    return obj if obj is not None else ""


_COLUMNS = [
    # fmt: off
    # Contact
    ("Nombre", lambda l: l.name),
    ("Teléfono", lambda l: l.phone),
    ("WhatsApp", lambda l: l.whatsapp_number or ("Sí" if l.has_whatsapp else "")),
    ("Emails", lambda l: "; ".join(l.emails_scraped or ([l.email] if l.email else []))),
    ("Sitio Web", lambda l: l.website),
    # Location
    ("Dirección", lambda l: l.address),
    ("Ciudad", None),
    # Ratings
    ("Rating Google", lambda l: l.rating),
    ("Reseñas", lambda l: l.reviews_count),
    # Digital
    ("Madurez Digital", lambda l: _get(l, "digital_maturity")),
    ("Stack Tech", lambda l: "; ".join(l.technology_stack or [])),
    (
        "Redes Sociales",
        lambda l: "; ".join(f"{k}: {v}" for k, v in (l.social_links or {}).items()),
    ),
    # Profile
    ("Tamaño Est.", lambda l: _get(l, "estimated_size")),
    ("Sector", lambda l: _get(l, "main_sector")),
    ("Tipo Comprador", lambda l: _get(l, "profile", "challenger_buyer_type")),
    ("Score Hormozi", lambda l: _get(l, "profile", "hormozi_score")),
    ("Label Hormozi", lambda l: _get(l, "profile", "hormozi_label")),
    ("Compromiso", lambda l: _get(l, "profile", "cardone_commitment")),
    ("Canal Entrada", lambda l: _get(l, "profile", "cardone_entry_channel")),
    ("Objeción Princ.", lambda l: _get(l, "profile", "cardone_objection")),
    ("Pitch Hook", lambda l: _get(l, "profile", "pitch_hook")),
    ("Acción Propuesta", lambda l: _get(l, "profile", "cardone_action_line")),
    # Timing
    ("Mejor Horario", lambda l: _get(l, "visit_timing", "timing_summary")),
    ("Conf. Horario", lambda l: _get(l, "visit_timing", "timing_confidence")),
    # Score + Tier
    ("Score Final", lambda l: round(l.final_score, 2) if l.final_score else ""),
    ("Tier", lambda l: l.tier),
    ("Prioridad", lambda l: l.contact_priority),
    ("Razón Descarte", lambda l: l.discard_reason or ""),
    # Source metadata
    ("Fuente", lambda l: l.source),
    ("Place ID", lambda l: l.place_id),
    # fmt: on
]


def _build_columns(city: str = "") -> list[tuple]:
    return [
        (header, extractor if extractor is not None else (lambda l, c=city: c))
        for header, extractor in _COLUMNS
    ]


# ──────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────


def export_to_excel(
    leads: Sequence[QualifiedLead],
    report: RunReport,
    output_dir: str = "output",
    filename_prefix: str = "prospectos",
    route_plan: RoutePlan | None = None,
    city: str = "",
) -> str:
    """
    Export leads + report to an Excel workbook.

    Returns:
        Absolute path of the created .xlsx file.
    """
    if not _OPENPYXL_OK:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = re.sub(r"[^\w]", "_", report.campaign_name.lower())
    out_path = Path(output_dir) / f"{filename_prefix}_{slug}_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    columns = _build_columns(city)

    # Tier sheets
    hot_leads = [l for l in leads if l.tier == "HOT"]
    warm_leads = [l for l in leads if l.tier == "WARM"]
    cold_leads = [l for l in leads if l.tier == "COLD"]

    _write_leads_sheet(wb, "🔴 HOT", hot_leads, "HOT", columns)
    _write_leads_sheet(wb, "🟠 WARM", warm_leads, "WARM", columns)
    _write_leads_sheet(wb, "🔵 COLD", cold_leads, "COLD", columns)
    _write_leads_sheet(wb, "📋 TODOS", list(leads), "TODOS", columns)
    if route_plan:
        _write_route_sheet(wb, route_plan)
    _write_summary_sheet(wb, report)

    wb.save(out_path)
    return str(out_path)


# ──────────────────────────────────────────────────────────────────
# Sheet writers
# ──────────────────────────────────────────────────────────────────


def _write_leads_sheet(
    wb: "Workbook",
    sheet_name: str,
    leads: list[QualifiedLead],
    tier_key: str,
    columns: list[tuple] | None = None,
) -> None:
    if columns is None:
        columns = _build_columns()
    ws = wb.create_sheet(title=sheet_name)
    header_fill = _HEADER_FILL.get(tier_key, _HEADER_FILL["TODOS"])
    row_fill = _TIER_ROW_FILL.get(tier_key)

    # Header row
    headers = [col[0] for col in columns]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = header_fill  # type: ignore[arg-type]
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER  # type: ignore[arg-type]

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, lead in enumerate(leads, start=2):
        fill = _TIER_ROW_FILL.get(lead.tier, row_fill)
        for col_idx, (_, extractor) in enumerate(columns, start=1):
            try:
                value = extractor(lead)
            except Exception:  # noqa: BLE001
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _BORDER  # type: ignore[arg-type]

    # Auto-fit columns (approximate)
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            ),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # Freeze header
    ws.freeze_panes = "A2"


def _write_summary_sheet(wb: "Workbook", report: RunReport) -> None:
    ws = wb.create_sheet(title="📊 RESUMEN")
    header_fill = _HEADER_FILL["RESUMEN"]

    rows = [
        ("Campo", "Valor"),
        ("Campaña", report.campaign_name),
        ("Fecha / Hora", report.timestamp),
        ("Leads raw captados", report.total_raw),
        ("Leads únicos (post-dedup)", report.total_after_dedup),
        ("HOT", report.hot_count),
        ("WARM", report.warm_count),
        ("COLD", report.cold_count),
        ("Duración (segundos)", report.duration_seconds),
        ("Iteraciones", report.iterations),
        ("Leads por iteración", ", ".join(str(n) for n in report.leads_per_iteration)),
    ]

    # Sources breakdown
    for source, count in report.sources_breakdown.items():
        rows.append((f"Fuente: {source}", count))

    # Error log
    if report.error_log:
        rows.append(("", ""))
        rows.append(("ERRORES", ""))
        for err in report.error_log:
            rows.append((err.get("step", "?"), err.get("message", "")))

    for row_idx, (key, val) in enumerate(rows, start=1):
        k_cell = ws.cell(row=row_idx, column=1, value=key)
        v_cell = ws.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            for cell in (k_cell, v_cell):
                cell.font = _HEADER_FONT
                cell.fill = header_fill  # type: ignore[arg-type]
        k_cell.border = _BORDER  # type: ignore[arg-type]
        v_cell.border = _BORDER  # type: ignore[arg-type]

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45


def _write_route_sheet(wb: "Workbook", route_plan: RoutePlan) -> None:
    ws = wb.create_sheet(title="RUTA")
    headers = [
        "Orden de Visita",
        "Negocio",
        "Dirección",
        "Teléfono",
        "Tier",
        "Score",
        "Distancia al siguiente",
        "Tiempo al siguiente",
        "Hora estimada de llegada",
        "Google Maps",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _ROUTE_HEADER_FONT
        cell.fill = _HEADER_FILL["RUTA"]  # type: ignore[arg-type]
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER  # type: ignore[arg-type]

    for row_idx, waypoint in enumerate(route_plan.waypoints, start=2):
        values = [
            waypoint.visit_order,
            waypoint.lead_name,
            waypoint.address,
            waypoint.phone,
            waypoint.tier,
            round(waypoint.final_score, 2),
            _format_distance_km(waypoint.distance_to_next_km),
            _format_duration_minutes(waypoint.duration_to_next_minutes),
            _format_eta(waypoint.estimated_arrival_minutes),
            "Abrir mapa",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _BORDER  # type: ignore[arg-type]
            if col_idx == 5:
                fill = _ROUTE_TIER_FILL.get(waypoint.tier)
                if fill:
                    cell.fill = fill
            if col_idx == 10 and waypoint.google_maps_url:
                cell.hyperlink = waypoint.google_maps_url
                cell.style = "Hyperlink"

    total_row = ws.max_row + 1
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = _BORDER  # type: ignore[arg-type]
        if col_idx in {1, 7, 8}:
            cell.fill = _HEADER_FILL["RUTA"]  # type: ignore[arg-type]

    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(
        row=total_row,
        column=7,
        value=_format_distance_km(route_plan.total_distance_km),
    )
    ws.cell(
        row=total_row,
        column=8,
        value=_format_duration_minutes(route_plan.total_duration_minutes),
    )

    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (
                len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
                for row_idx in range(1, ws.max_row + 1)
            ),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"


def _format_distance_km(distance_km: float) -> str:
    if distance_km <= 0:
        return ""
    return f"{distance_km:.2f} km"


def _format_duration_minutes(duration_minutes: float) -> str:
    if duration_minutes <= 0:
        return ""
    rounded = int(round(duration_minutes))
    hours, minutes = divmod(rounded, 60)
    if hours:
        return f"{hours}h {minutes}min"
    return f"{minutes} min"


def _format_eta(minutes_from_start: float) -> str:
    rounded = int(round(minutes_from_start))
    hours, minutes = divmod(rounded, 60)
    if hours:
        return f"T+{hours}h {minutes}min"
    return f"T+{minutes} min"
    ws.freeze_panes = "A2"
